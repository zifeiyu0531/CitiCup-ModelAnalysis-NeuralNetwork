########################
#
# 模型调用过程
# 2019/7/1 v1.0
#
########################
from gensim.models import word2vec
import openpyxl
import gensim
import jieba
from target import Target
from xlutils.copy import copy

global MODEL_NAME, TARGET_FILE_NAME, TARGET_LIST, WORD_LIST
MODEL_NAME = 'models\CitiCup_1000.model.bin'            # 模型名
TARGET_FILE_NAME = 'test\\600054_3.txt'       # 待分析文件名
WORD_LIST = []                          # 待分析词列表
TARGET_LIST = [
    [
        Target('1公司年报是否披露主要的财务指标', [
            ['资产', '负债率'],
            ['流动', '比率'],
            ['速动', '比率'],
            ['应收', '账款', '周转率'],
            ['存货', '周转率'],
            ['资本金', '利润率'],
            ['销售', '利润率'],
            ['营业', '收入', '利润率'],
            ['成本', '费用', '利润率']
        ], 'struct'),
        Target('1公司年报是否披露主要的财务指标', [
            ['资产', '负债率'],
            ['流动', '比率'],
            ['速动', '比率'],
            ['应收', '账款', '周转率'],
            ['存货', '周转率'],
            ['资本金', '利润率'],
            ['销售', '利润率'],
            ['营业', '收入', '利润率'],
            ['成本', '费用', '利润率']
        ], 'struct'),
        Target('2公司年报是否披露其他应收、应付款等带"其他"字样的项目的详细情况', [
            ['其他', '应收款'],
            ['其他', '应付款'],
            ['赔款'],
            ['罚款'],
            ['出租', '包装物', '租金'],
            ['职工', '水电费'],
            ['职工', '医药费'],
            ['租入', '包装物', '押金'],
            ['暂付']
        ], 'struct'),
        Target('3公司是否披露营业外收支的详细情况', [
            ['罚款', '支出'],
            ['捐赠', '支出'],
            ['非常', '损失']
        ], 'struct'),
        Target('4公司是否披露披露募集资金有关项目的进展', [
            ['合同', '完工', '进度'],
            ['累计', '实际', '发生']
        ], 'struct'),
        Target('5投资项目未达到计划进度和收益的，公司是否披露原因', [
            ['项目', '受挫', '原因'],
            ['项目', '失败', '原因'],
            ['项目', '未到', '计划', '原因']
        ], 'union'),
        Target('6公司募集资金投资项目变更的是否披露变更原因', [
            ['投资', '项目', '变更', '原因'],
            ['放弃', '项目']
        ], 'union'),
        Target('7公司年报是否说明关联交易的必要性、持续性', [
            ['关联', '交易'],
            ['经营', '租赁'],
            ['融资', '租赁'],
            ['代理', '销售'],
            ['代理', '签订', '合同'],
            ['研究', '开发', '转移'],
            ['许可', '协议']
        ], 'union'),
        Target('8公司若不存在重大诉讼、仲裁事项，是否在年报中明确声明', [
            ['没有', '诉讼'],
            ['没有', '仲裁'],
            ['申诉', '仲裁'],
            ['权利', '仲裁'],
            ['利益', '仲裁'],
            ['合约', '仲裁']
        ], 'union'),
        Target('9公司是否披露支付的审计费用金额', [
            ['审计','费用']
        ], 'struct'),
        Target('10若不存在任何重大担保合同，公司是否在年报中声明', [
            ['没有', '担保', '合同']
        ], 'struct'),
        Target('11公司年报是否披露更换会计事务所的原因', [
            ['会计', '事务所', '变更', '原因']
        ], 'struct'),
        Target('12公司是否披露社会责任报告书', [
            ['社会', '责任', '报告']
        ], 'struct'),
        Target('13年报中是否披露公司建有网站', [
            ['本公司', '网址']
        ], 'struct'),
        Target('14公司年报中是否披露重要的会计政策', [
            ['会计', '政策']
        ], 'struct'),
        Target('15是否及时披露董事会、监事会和股东大会的决议报告', [
            ['决议', '日期']
        ], 'struct')
    ],
    [
        Target('1是否未发生会计变更', [
            ['未发生', '会计', '政策', '变更'],
            ['未发生', '会计', '估计', '变更'],
            ['未发生', '会计', '差错', '更正']
        ], 'union'),
        Target('2是否披露发生会计变更原因', [
            ['会计', '变更', '原因']
        ], 'struct')
    ],
    [
        Target('1两职分离能有效地提高监督和控制经理层的能力、维护董事会的独立性、保证会计信息披露质量。', [
            ['两职', '分离'],
            ['两权', '分离']
        ], 'union'),
        Target('2是否有独立董事', [
            ['独立', '董事']
        ], 'struct'),
        Target('3是否有审计委员会', [
            ['审计', '委员会']
        ], 'struct'),
        Target('4公司是否披露内部控制', [
            ['组织', '控制'],
            ['人员', '控制'],
            ['职务', '控制'],
            ['业务', '控制'],
            ['授权', '控制']
        ], 'struct')
    ],
    [
        Target('1公司是否根据风险情况，披露拟采取的对策', [
            ['风险', '预防']
        ], 'struct'),
        Target('2公司是否自愿披露经审核的新年度盈利预测', [
            ['盈利', '预测']
        ], 'struct'),
        Target('3公司年报是否披露新年度的经营计划或经营目标', [
            ['经营', '目标']
        ], 'struct'),
        Target('4公司是否对公司战略进行描述', [
            ['战略']
        ], 'struct'),
        Target('5公司是否进行行业未来的发展趋势分析', [
            ['公司', '未来'],
            ['行业', '未来']
        ], 'union'),
        Target('6公司是否对面临的市场竞争格局进行分析', [
            ['市场', '格局'],
            ['市场', '规模'],
            ['市场', '分析'],
            ['竞争', '格局']
        ], 'union'),
        Target('7公司是否对未来公司发展机遇和挑战进行分析', [
            ['机会', '挑战']
        ], 'struct'),
        Target('8公司是否对面临的风险因素进行分析', [
            ['面临', '风险'],
            ['潜在', '威胁']
        ], 'union')
    ]
]


##########################################
# test_step()
# 验证过程入口函数
##########################################
def test_step():
    cut_txt()
    check_grade()


##########################################
# cut_txt()
# 将待验证文件分词，存入WORD_LIST
##########################################
def cut_txt():
    global TARGET_FILE_NAME, WORD_LIST
    try:
        old_file = open(TARGET_FILE_NAME, 'r', encoding='utf-8')
        section_list = old_file.read().split('$')               # 分段
        for section in section_list:                            # 将段落分词，存入WORD_LIST
            cut_text = jieba.cut(section, cut_all=False)
            new_text = ' '.join(cut_text)
            new_list = new_text.split(' ')
            WORD_LIST.append(new_list)
        old_file.close()
    except BaseException as e:
        print(Exception, ":", e)


##########################################
# check_grade()
# 相似度检测
##########################################
def check_grade():
    global MODEL_NAME, TARGET_LIST, WORD_LIST
    xlsx_file = openpyxl.load_workbook('output\model_1000.xlsx')         # 获取表格文件
    model = gensim.models.KeyedVectors.load_word2vec_format(MODEL_NAME, binary=True)          # 加载已训练好的模型

    for index, lists in enumerate(WORD_LIST):
        for target_index, sub_target_list in enumerate(TARGET_LIST):  # 计算两个词的相似度/相关程度
            for sub_target_index, target in enumerate(sub_target_list):
                xlsx_file.worksheets[target_index].cell(1, sub_target_index+2, target.name)
                if target.type == "struct":
                    for value_list in target.value_list:
                        time_list = []
                        for key_word in value_list:
                            time = 0
                            for word in lists:
                                try:
                                    grade = model.wv.similarity(key_word, word)
                                    if grade > 0.7:
                                        time += 1
                                except KeyError:
                                    pass
                            time_list.append(time)
                        print("指标" + str(value_list) + "在第" + str(index + 1) + "段中匹配的次数为" + str(time_list))
                        old_cell_value = str(xlsx_file.worksheets[target_index].cell(row=index + 2, column=sub_target_index + 2).value)
                        if old_cell_value != 'None':
                            xlsx_file.worksheets[target_index].cell(index + 2, sub_target_index + 2, old_cell_value + str(time_list))
                        else:
                            xlsx_file.worksheets[target_index].cell(index + 2, sub_target_index + 2, str(time_list))
                else:
                    for value_list in target.value_list:
                        time_list = []
                        for key_word in value_list:
                            time = 0
                            for word in lists:
                                try:
                                    grade = model.wv.similarity(key_word, word)
                                    if grade > 0.7:
                                        time += 1
                                except KeyError:
                                    pass
                            time_list.append(time)
                        print("指标" + str(value_list) + "在第" + str(index + 1) + "段中匹配的次数为" + str(time_list))
                        old_cell_value = str(xlsx_file.worksheets[target_index].cell(row=index + 2, column=sub_target_index + 2).value)
                        if old_cell_value != 'None':
                            xlsx_file.worksheets[target_index].cell(index + 2, sub_target_index + 2, old_cell_value + str(time_list))
                        else:
                            xlsx_file.worksheets[target_index].cell(index + 2, sub_target_index + 2, str(time_list))

    xlsx_file.save(filename='output\model_1000.xlsx')


def main():
    test_step()


if __name__ == '__main__':
    main()
